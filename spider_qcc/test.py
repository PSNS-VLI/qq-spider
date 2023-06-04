import sys
import io
import json
import openpyxl
from agent import XlsxAgent
from calculator import TextCalculator as TC
sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030')

# wb = openpyxl.load_workbook('data/excel_output/企业名单(3787家)(2).xlsx', read_only = True)
# ws = wb.active

# table = []
# pre_data = []
# error_data = []
# r_i = 0
# for row in ws.rows:
#     item = []
#     for col in row:
#         item.append(col.value)
#     table.append(item)
#     if row[0].value == None:
#         error_data.append({'index': r_i, 'company': row[1].value})
#     r_i += 1

# with open('data/temp_json_data/2023-02-22@21-55@pre_data.json', 'r', encoding='utf-8') as f:
#     pre_data = json.loads(f.read())['data']

# for item in error_data:
#     print(item['company'], pre_data[item['index']-1]['company'], sep="\n")
#     possible = pre_data[item['index']-1]['BING'] + pre_data[item['index']-1]['BAIDU']
#     if len(possible) > 0:
#         possible =  TC.sort_dict_by_similar(possible, 'title')[0]['item']
#         table[item['index']][0] = possible['title']

# wb = openpyxl.Workbook()
# ws = wb.create_sheet()
# for r in range(len(table)):
#     for c in range(len(table[r])):
#         cell = ws.cell(r+1, c+1, table[r][c])
# wb.save('data/excel_output/企业名单.xlsx')

wb = openpyxl.load_workbook('data/excel_output/缺失名单523(2).xlsx', read_only = True)
ws = wb.active

table = []
pre_data = []
error_data = []
r_i = 0
for row in ws.rows:
    item = []
    if row[0].value == None:
        for col in row:
            item.append(col.value)
        error_data.append(item)
    else:
        for col in row:
            item.append(col.value)
        table.append(item)
    r_i += 1

wb = openpyxl.Workbook()
ws = wb.create_sheet()
for r in range(len(table)):
    for c in range(len(table[r])):
        cell = ws.cell(r+1, c+1, table[r][c])
wb.save(f'data/excel_output/企业名单{len(table)-1}.xlsx')

wb = openpyxl.Workbook()
ws = wb.create_sheet()
for r in range(len(error_data)):
    for c in range(len(error_data[r])):
        cell = ws.cell(r+1, c+1, error_data[r][c])
wb.save(f'data/excel_output/缺失名单{len(error_data)-1}.xlsx')