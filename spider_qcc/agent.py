import xlrd
import xlwt
import openpyxl
from abc import ABCMeta, abstractmethod

class Agent:
    default_save_path = r'D:\\'
    def __init__(self, filename):
        self.filename = filename

    def __str__(self):
        print(f'<{self.__class__.__name__} {self.filename}>')

    @abstractmethod
    def read_data(self):
        pass

    @abstractmethod
    def write_data(self):
        pass

    @abstractmethod
    def save_as(self):
        pass

    def generate_file_name(self, _name, _path):
        _name_l = (_name or self.filename.split('/')[-1:][0]).split('.')
        _name_l[0] = _name_l[0] + '(2)'
        _name = '.'.join(_name_l)
        return f'{_path or self.default_save_path}{_name}'

    def set_scope(self, start_row: int, start_col: int,
        end_row: int, end_col: int):
        self.start_row = start_row
        self.start_col = start_col
        self.end_row = end_row
        self.end_col = end_col

    def get_scope(self) -> tuple:
        
        s_r = self.start_row if self.start_row else None
        e_r = self.end_row if self.end_row else None
        s_c = self.start_col if self.start_col else None
        e_c = self.end_col if self.end_col else None
        return (s_r, e_r, s_c, e_c)

class ExcelAgent(Agent):

    def __init__(self, filename: str):
        super().__init__(filename)
        self.__file = None
        self.__init(filename)

    def read_data(self, sheet_index: int = 0, sheet_name: int = None) -> list:
        table = self.__get_sheet(sheet_index = sheet_index,
            sheet_name = sheet_name)
        data = []

        for i in range( self.start_row or 0, self.end_row or table.nrows):
            data.append(table.row_values(i))
        return data

    def read_data_with_head(self, sheet_index: int = 0, sheet_name: int = None,
        head_index: int = 0) -> dict:
        table = self.__get_sheet(sheet_index = sheet_index,
            sheet_name = sheet_name)
        data = {
            'head': table.row_values(head_index)
        }
        _data = []

        for i in range(self.start_row  or 1, self.end_row or table.nrows):
            if i == head_index:
                continue
            _data.append(table.row_values(i))
        data['data'] = _data
        return data


    def write_data(self, data: list, output_name: str = None,
        output_path: str = None):
        sheet = self.__set_sheet('sheet_new')

        for r in range(len(data)):
            for c in range(len(data[r])):
                sheet.write(r, c, data[r][c])

        sheet.get_parent().save(self.generate_file_name(output_name,
            output_path))

    def save_as(self, data: list, output_name: str = None,
        output_path: str = None):
        self.write_data(data, output_name, output_path)

    def __init(self, filename):
        try:
            self.__file = self.__get_file(filename)
        except Exception as result:
            print(result)
            raise IOError('打开文件失败')

    def __get_sheet(self, sheet_index: int = 0,
        sheet_name: int = None):
        return self.__file.sheet_by_name(sheet_name) if sheet_name else\
            self.__file.sheet_by_index(sheet_index)

    def __set_sheet(self, sheet_name: str):
        self.__new_book_sheet(sheet_name)

    def __new_book_sheet(self, sheet_name: str):
        return xlwt.Workbook(encoding='utf-8').add_sheet(sheet_name)

    def __get_file(self, filename: str):
        self.__file = self.__file or xlrd.open_workbook(filename)
        return self.__file

class XlsxAgent(Agent):
    __max_col = None
    __max_row = None
    __valid_col = None
    __valid_row = None
    __file = None
    def __init__(self, filename: str):
        super().__init__(filename)
        self.__file = None
        self.__init(filename)

    def read_data(self, sheet_index: int = 0, sheet_name: int = None) -> list:
        table = self.__get_sheet(sheet_index = sheet_index,
            sheet_name = sheet_name)
        max_row = self.end_row or self.__get_max_row(table)
        max_col = self.end_col or self.__get_max_col(table)
        data = []

        for r in range(self.start_row or 1, max_row):
            item = []
            for c in range(self.start_col or 1, max_col):
                item.append(table.cell(row = r, column = c).value)
            data.append(item)
        return data

    def read_data_with_head(self, sheet_index: int = 0, sheet_name: int = None,
        head_index: int = 1) -> dict:
        table = self.__get_sheet(sheet_index = sheet_index,
            sheet_name = sheet_name)
        max_row = self.end_row or self.__get_max_row(table)
        max_col = self.end_col or self.__get_max_col(table)
        data = {}
        _data = []
        _head = []

        r_i = 1
        c_i = 1
        for row in table.rows:
            if r_i < (self.start_row or 1):
                r_i = r_i + 1
                continue
            elif r_i > max_row:
                break
            elif r_i == head_index:
                for col in row:
                    if c_i < (self.start_col or 1):
                        c_i = c_i + 1
                        continue
                    elif c_i > max_col:
                        break
                    else:
                        _head.append(col.value)
                r_i = r_i + 1
            else:
                print(f'number {r_i}', flush=True)
                item = []
                for col in row:
                    if c_i < (self.start_col or 1):
                        c_i = c_i + 1
                        continue
                    elif c_i > max_col:
                        break
                    else:
                        item.append(col.value)
                _data.append(item)
                r_i = r_i + 1

        # for r in range(self.start_row or 1, max_row):
        #     print(f'number {r}', flush=True)
        #     if r == head_index:
        #         for c in range(self.start_col or 1, max_col):
        #             _head.append(table.cell(row = r, column = c).value)
        #     else:
        #         item = []
        #         for c in range(self.start_col or 1, max_col):
        #             item.append(table.cell(row = r, column = c).value)
        #         _data.append(item)
        data['head'] = _head
        data['data'] = _data
        return data


    def write_data(self, data: list, output_name: str = None,
        output_path: str = None, factory: '工厂函数' = None):
        sheet = self.__set_sheet('sheet_new')

        for r in range(1, len(data)+1):
            for c in range(1, len(data[r])+1):
                cell = sheet.cell(r, c, data[r][c])
                if factory:
                    factory(cell)

        self.__file.save(self.generate_file_name(output_name,
            output_path))

    def save_as(self, data: list, output_name: str = None,
        output_path: str = None, factory: '工厂函数' = None):
        book, sheet = self.__new_book_sheet('sheet_new')
        for r in range(len(data)):
            for c in range(len(data[r])):
                cell = sheet.cell(r+1, c+1, data[r][c])
                if factory:
                    factory(cell)

        book.save(self.generate_file_name(output_name,
            output_path))

    def __init(self, filename):
        try:
            self.__file = self.__get_file(filename)
        except Exception as result:
            print(result)
            raise IOError('打开文件失败')

    @staticmethod
    def bg_color_factory(cell, color: str = 'FF0000'):
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=color)

    def __get_sheet(self, sheet_index: int = 0,
        sheet_name: int = None):
        return self.__file.get_sheet_by_name(sheet_name) if sheet_name else\
            self.__file.worksheets[sheet_index]

    def __set_sheet(self, sheet_name: str = 'sheet1'):
        return  self.__file.create_sheet(sheet_name)

    def __new_book_sheet(self, sheet_name: str):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(sheet_name)
        return (wb, ws)

    def __get_file(self, filename: str):
        self.__file = self.__file or openpyxl.load_workbook(filename, read_only = True)
        return self.__file

    def __get_max_row(self, sheet):
        self.__max_row = self.__max_row or self.__get_max(sheet, 'row')
        return self.__max_row

    def __get_max_col(self, sheet):
        self.__max_col = self.__max_col or self.__get_max(sheet, 'col')
        return self.__max_col

    def __get_valid_row(self, sheet):
        self.__valid_row = self.__valid_row or self.__valid_max(sheet, 'row')
        return self.__valid_row

    def __get_valid_col(self, sheet):
        self.__valid_col = self.__valid_col or self.__get_valid(sheet, 'col')
        return self.__valid_col

    def __get_max(self, sheet, tag: str):
        i = eval(f'sheet.max_{tag}')
        rows = eval(f'sheet.{tag}s')
        real_max = 0
        while i > 0:
            if tuple(rows)[i-1] == None:
                i = i - 1
            else:
                real_max = i
                break
        return real_max

    def __get_valid(self, sheet, tag: str):
        i = eval(f'sheet.max_{tag}')
        rows = eval(f'sheet.{tag}s')
        _max = eval(f'self.__get_max_{tag}()')
        valid = 0
        while i < _max:
            if tuple(rows)[i-1]:
                valid = valid + 1
            i = i - 1
                
        return valid
